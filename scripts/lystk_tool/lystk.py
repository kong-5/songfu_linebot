"""
lystk — LyERP API tool for 龍港集團 (松富/松揚/松成/龍港).

USAGE AS LIBRARY (any future script):

    import sys
    sys.path.insert(0, r"D:\\Work\\lystk_tool")
    import lystk

    # one-time setup (interactive)
    lystk.setup_credentials()

    # check connection
    print(lystk.version())

    # query one day's sales bills for 松富
    records = lystk.query(icpno="00", idakd="0000A1", date="yesterday")
    for r in records:
        print(r["SP_NO"], r["SP_CTNAME"], r["SP_TOTAL"])

    # use friendly names
    records = lystk.query(company="松富", kind="銷貨單", date="2026-05-19")

USAGE AS CLI:

    py D:\\Work\\lystk_tool\\lystk.py setup
    py D:\\Work\\lystk_tool\\lystk.py version
    py D:\\Work\\lystk_tool\\lystk.py companies
    py D:\\Work\\lystk_tool\\lystk.py kinds
    py D:\\Work\\lystk_tool\\lystk.py query 00 0000A1 yesterday
    py D:\\Work\\lystk_tool\\lystk.py query 松富 銷貨單 today
    py D:\\Work\\lystk_tool\\lystk.py dump 松富 銷貨單 2026-05-19 D:\\out.json
    py D:\\Work\\lystk_tool\\lystk.py dump 松富 銷貨單 2026-05-19 D:\\out.xlsx
"""
import sys, os, time, json, datetime

# ============================================================
#  Constants — server, credential storage, code maps
# ============================================================

API_URL = "http://192.168.4.11/erpapi/erpservice.svc?wsdl"
KEYRING_SERVICE = "LyERPApi"

# Company ICPNO ↔ name
COMPANIES = {
    "00": "松富物流股份有限公司",
    "01": "龍港",
    "02": "松揚",
    "03": "松成物流股份有限公司",
}
# Reverse name → ICPNO (handle short/full names)
_NAME_TO_ICPNO = {
    "松富": "00", "松富物流": "00", "松富物流股份有限公司": "00",
    "龍港": "01",
    "松揚": "02", "松揚物流": "02", "松揚物流股份有限公司": "02",
    "松成": "03", "松成物流": "03", "松成物流股份有限公司": "03",
}

# Data kind (idakd) ↔ name
IDAKD = {
    "000000": "貨品基本資料",
    "000001": "客戶基本資料(寫入)",
    "000004": "倉庫基本資料",
    "000009": "目前庫存(廠內倉)",
    "00000D": "客戶基本資料(讀取)",
    "0000A0": "訂貨單",
    "0000A1": "銷貨單",
    "0000A2": "銷貨退回單",
}
_KIND_TO_IDAKD = {v: k for k, v in IDAKD.items()}
# Aliases
_KIND_TO_IDAKD.update({
    "銷貨": "0000A1", "銷退": "0000A2", "銷退貨": "0000A2",
    "訂單": "0000A0", "訂貨": "0000A0",
    "客戶": "00000D", "貨品": "000000", "倉庫": "000004", "庫存": "000009",
})

# Schema field prefix per idakd — used for date filtering, etc
_FIELD_PREFIX = {
    "0000A0": "OR",   # 訂單 OR_*
    "0000A1": "SP",   # 銷貨 SP_*
    "0000A2": "SP",   # 銷退 SP_*
    "000000": "SK",   # 貨品 SK_*
    "00000D": "CT",   # 客戶 CT_*
    "000001": "CT",   # 客戶 CT_*
    "000004": "WH",   # 倉庫 (guess)
    "000009": "ST",   # 庫存 (guess)
}

# ============================================================
#  Credential management
# ============================================================

def _get_keyring():
    try:
        import keyring
    except ImportError:
        raise RuntimeError("Missing dependency: pip install keyring")
    return keyring

def get_credentials():
    """Return (pusid, pverifykey) tuple. Raises if not set."""
    k = _get_keyring()
    pusid = k.get_password(KEYRING_SERVICE, "pusid")
    pvk   = k.get_password(KEYRING_SERVICE, "pverifykey")
    if not pusid or not pvk:
        raise RuntimeError(
            "Credentials not stored. Run: py lystk.py setup\n"
            "or call lystk.setup_credentials() interactively."
        )
    return pusid, pvk

def setup_credentials(pusid=None, pverifykey=None):
    """
    Store API credentials in Windows Credential Manager (per-user encrypted).
    If args are None, prompts interactively (password hidden).
    """
    import getpass
    if pusid is None:
        pusid = input("API account (pusid): ").strip()
    if not pusid: raise ValueError("pusid is empty")
    if pverifykey is None:
        pverifykey = getpass.getpass("API password (pverifykey, hidden): ")
    if not pverifykey: raise ValueError("pverifykey is empty")

    # Test before storing
    print("Testing connection...")
    client = _build_client()
    key = client.service.LyGetPassKey(pusid=pusid, pverifykey=pverifykey)
    try:
        if int(key) < 0:
            errs = {-1:"SQL 連接失敗", -2:"帳號不存在", -3:"密碼不符", -4:"帳號存在沒有權限"}
            raise RuntimeError(f"Auth failed: {errs.get(int(key), 'unknown')} (code={key})")
    except (ValueError, TypeError):
        pass
    print(f"  [OK] auth successful (test key = {key})")

    k = _get_keyring()
    k.set_password(KEYRING_SERVICE, "pusid", pusid)
    k.set_password(KEYRING_SERVICE, "pverifykey", pverifykey)
    print("  [OK] credentials stored in Windows Credential Manager.")

# ============================================================
#  SOAP client (zeep)
# ============================================================

_client = None
def _build_client():
    try:
        from zeep import Client, Settings
    except ImportError:
        raise RuntimeError("Missing dependency: pip install zeep")
    return Client(API_URL, settings=Settings(strict=False, xml_huge_tree=True))

def get_client():
    global _client
    if _client is None:
        _client = _build_client()
    return _client

def version():
    """Return API version string. No auth needed."""
    return get_client().service.LyGetERPApiVer()

def fresh_key(retries=3):
    """Get a fresh 30-second API key. Auto-retries on transient errors."""
    pusid, pvk = get_credentials()
    client = get_client()
    last_err = None
    for i in range(retries):
        try:
            k = client.service.LyGetPassKey(pusid=pusid, pverifykey=pvk)
            try:
                if int(k) < 0:
                    raise RuntimeError(f"LyGetPassKey returned {k}")
            except (ValueError, TypeError):
                pass
            return k
        except Exception as e:
            last_err = e
            time.sleep(1 + i)
    raise RuntimeError(f"LyGetPassKey failed after {retries} tries: {last_err}")

# ============================================================
#  Argument coercion (friendly names → codes)
# ============================================================

def resolve_icpno(arg):
    """Accept '00' / '松富' / etc → '00'."""
    arg = str(arg).strip()
    if arg in COMPANIES: return arg
    if arg in _NAME_TO_ICPNO: return _NAME_TO_ICPNO[arg]
    raise ValueError(f"Unknown company: {arg!r}. Try one of {list(COMPANIES)} or {list(_NAME_TO_ICPNO)}")

def resolve_idakd(arg):
    """Accept '0000A1' / '銷貨單' / '銷貨' / etc → '0000A1'."""
    arg = str(arg).strip()
    if arg in IDAKD: return arg
    if arg in _KIND_TO_IDAKD: return _KIND_TO_IDAKD[arg]
    raise ValueError(f"Unknown kind: {arg!r}. Try one of {list(IDAKD)} or {list(_KIND_TO_IDAKD)}")

def resolve_date(arg):
    """Accept 'today' / 'yesterday' / YYYY-MM-DD / YYYYMMDD → datetime.date."""
    if isinstance(arg, datetime.date): return arg
    s = str(arg).strip().lower()
    today = datetime.date.today()
    if s in ("today", "t"): return today
    if s in ("yesterday", "y"): return today - datetime.timedelta(days=1)
    if s in ("tomorrow",):     return today + datetime.timedelta(days=1)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try: return datetime.datetime.strptime(s, fmt).date()
        except: continue
    raise ValueError(f"Bad date: {arg!r}")

# ============================================================
#  Main query function
# ============================================================

def query(icpno=None, idakd=None, date=None, *,
          company=None, kind=None,
          start=None, end=None,
          where="", whval="", order=None,
          limit=0):
    """
    Query records from LyERP API.

    Args:
      icpno    : '00' / '03' / etc. OR use company='松富'.
      idakd    : '0000A1' / etc. OR use kind='銷貨單'.
      date     : single day. OR use start + end for a range.
      where    : extra SQL WHERE clause with '@v1@', '@v2@' placeholders.
      whval    : values for the placeholders, separated by '@#1#@'.
      order    : ORDER BY clause (default sorts by main NO field).
      limit    : 0 = no pagination, >0 = page size.

    Returns:
      list of dicts (one per record), with all fields from API.
    """
    icpno = resolve_icpno(company or icpno)
    idakd = resolve_idakd(kind or idakd)

    # Date handling
    prefix = _FIELD_PREFIX.get(idakd, "SP")
    date_field = f"{prefix}_DATE" if prefix in ("SP",) else (
        f"{prefix}_DATE1" if prefix == "OR" else None
    )

    user_where = where
    user_whval = whval

    if date is not None and not (start or end):
        d = resolve_date(date)
        if date_field is None:
            raise ValueError(f"date filter not supported for idakd={idakd}; use where=")
        date_where = f"{date_field} between '@v1@' and '@v2@'"
        date_whval = f"{d.strftime('%Y-%m-%d')} @#1#@ {d.strftime('%Y-%m-%d')} 23:59:59"
        if user_where:
            user_where = f"({user_where}) and ({date_where})"
            user_whval = f"{user_whval} @#1#@ {date_whval.split(' @#1#@ ')[0]} @#1#@ {date_whval.split(' @#1#@ ')[1]}"
        else:
            user_where = date_where
            user_whval = date_whval
    elif start or end:
        sd = resolve_date(start) if start else resolve_date(end)
        ed = resolve_date(end) if end else resolve_date(start)
        if date_field is None:
            raise ValueError(f"date range not supported for idakd={idakd}")
        range_where = f"{date_field} between '@v1@' and '@v2@'"
        range_whval = f"{sd.strftime('%Y-%m-%d')} @#1#@ {ed.strftime('%Y-%m-%d')} 23:59:59"
        if user_where:
            user_where = f"({user_where}) and ({range_where})"
            user_whval = f"{user_whval} @#1#@ {range_whval}"
        else:
            user_where = range_where
            user_whval = range_whval

    # Default ordering
    if order is None:
        no_field = f"{prefix}_NO"
        order = f"order by {no_field}"

    client = get_client()
    resp = client.service.LyDataOut(
        ikye=fresh_key(),
        icpno=icpno, idakd=idakd,
        ifld="", idetfields="",
        irwhere=user_where, iwhval=user_whval,
        irec=int(limit), imode=" " * 30,
        iorder=order, idtorder="",
        iswhere="", isifld="",
        Isecgroup="", iseckindfg="", iseckind="",
        Isecorder="", Isecrec=0,
    )
    rc = str(resp["LyDataOutResult"])
    if rc != "0":
        errs = {"-1":"SQL 連接失敗","-2":"讀取失敗","-3":"金鑰失效",
                "-4":"金鑰不合法","-5":"無權限"}
        raise RuntimeError(f"LyDataOut failed: {errs.get(rc, f'code={rc}')}")
    xml = resp["ixmlda"]
    if not xml:
        return []
    from xml.etree import ElementTree as ET
    root = ET.fromstring(str(xml))
    rows = []
    for t in root.findall(".//LYDATATITLE"):
        d = {}
        for child in t:
            d[child.tag] = (child.text or "").strip() if child.text else ""
        rows.append(d)
    return rows

# ============================================================
#  Dump helpers
# ============================================================

def dump_json(records, out_path):
    out_path = os.path.abspath(out_path)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    return out_path

def dump_xlsx(records, out_path):
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("Missing dependency: pip install openpyxl")
    out_path = os.path.abspath(out_path)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    if not records:
        ws["A1"] = "(no records)"
    else:
        cols = list(records[0].keys())
        # union of keys (in case rows have different fields)
        for r in records[1:]:
            for k in r.keys():
                if k not in cols: cols.append(k)
        for c, name in enumerate(cols, 1):
            ws.cell(1, c, name).font = openpyxl.styles.Font(bold=True)
        for ri, r in enumerate(records, 2):
            for c, name in enumerate(cols, 1):
                ws.cell(ri, c, r.get(name, ""))
    wb.save(out_path)
    return out_path

# ============================================================
#  CLI
# ============================================================

def _cli_companies():
    print("ICPNO  Name")
    for k, v in COMPANIES.items():
        print(f"  {k}     {v}")

def _cli_kinds():
    print("idakd   名稱")
    for k, v in IDAKD.items():
        print(f"  {k}  {v}")

def _cli_query(argv):
    if len(argv) < 3:
        print("Usage: query <icpno|company> <idakd|kind> <date>"); return 1
    icpno = resolve_icpno(argv[0])
    idakd = resolve_idakd(argv[1])
    date  = resolve_date(argv[2])
    rows = query(icpno=icpno, idakd=idakd, date=date)
    print(f"{len(rows)} record(s) for {COMPANIES.get(icpno, icpno)} / {IDAKD.get(idakd, idakd)} / {date}")
    if not rows: return 0
    # Print key fields based on prefix
    prefix = _FIELD_PREFIX.get(idakd, "SP")
    keys = {
        "SP": ["SP_NO", "SP_DATE", "SP_CTNO", "SP_CTNAME", "SP_TOTAL"],
        "OR": ["OR_NO", "OR_DATE1", "OR_CTNO", "OR_CTNAME", "OR_SUM"],
        "CT": ["CT_NO", "CT_NAME", "CT_SNAME", "CT_ADDR1", "CT_TEL"],
        "SK": ["SK_NO", "SK_NAME", "SK_SPEC", "SK_UNIT"],
    }.get(prefix, list(rows[0].keys())[:6])
    print()
    print("  " + "  ".join(f"{k:<20}" for k in keys))
    print("  " + "  ".join("-" * 20 for _ in keys))
    for r in rows:
        print("  " + "  ".join(f"{str(r.get(k, ''))[:20]:<20}" for k in keys))
    return 0

def _cli_dump(argv):
    if len(argv) < 4:
        print("Usage: dump <icpno|company> <idakd|kind> <date> <out.json|out.xlsx>"); return 1
    icpno = resolve_icpno(argv[0])
    idakd = resolve_idakd(argv[1])
    date  = resolve_date(argv[2])
    out   = argv[3]
    rows = query(icpno=icpno, idakd=idakd, date=date)
    if out.lower().endswith(".xlsx"):
        p = dump_xlsx(rows, out)
    else:
        p = dump_json(rows, out)
    print(f"saved {len(rows)} record(s) -> {p}")
    return 0

def _cli():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    argv = sys.argv[1:]
    if not argv:
        print(__doc__); return 0
    cmd = argv[0].lower()
    rest = argv[1:]
    try:
        if cmd in ("setup", "setup-credentials"):
            setup_credentials(); return 0
        if cmd in ("version", "ver"):
            print(version()); return 0
        if cmd in ("companies", "company", "icpno"):
            _cli_companies(); return 0
        if cmd in ("kinds", "kind", "idakd", "types"):
            _cli_kinds(); return 0
        if cmd == "query":
            return _cli_query(rest)
        if cmd == "dump":
            return _cli_dump(rest)
        print(f"Unknown command: {cmd}\n{__doc__}")
        return 1
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return 2

if __name__ == "__main__":
    sys.exit(_cli())
